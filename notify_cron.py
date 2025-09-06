import time
import random
from dotenv import find_dotenv, load_dotenv
import os
import time
from datetime import datetime, timedelta
import re
import pandas as pd
import numpy as np
import time
from PIL import Image
import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

from bhavcopy_login import login
from bhavcopy_utils import *
# from queries import get_mail_template, get_holidays_for_year, get_exception_trading_dates_to_year, get_expry_days
from utils import (save_bhav_copy_data, create_or_add_master_data, get_google_script_data)
from corp_actions import adjust_corp_actions
from financial_analyser import FinancialDataAnalyzer

dotenv_path = find_dotenv()

if dotenv_path:
    load_dotenv(dotenv_path=dotenv_path, override=True)
else:
    print("No .env file found")

NSE_base_url = os.getenv('NSE_base_url')
SENDER_EMAIL = os.getenv('SENDER_EMAIL')
SENDER_PASSWORD = os.getenv('SENDER_PASSWORD')
SMTP_SERVER = os.getenv('SMTP_SERVER')
SMTP_PORT = int(os.getenv('SMTP_PORT'))
BHAVCOPY_FILE_NAMES = ['F&O-UDiFF Common Bhavcopy Final (zip)', 'F&O-Participant wise Open Interest (csv)', 'Full Bhavcopy and Security Deliverable data']
MAIL_RECIPIENTS = ['goela.shubh@gmail.com']
MAIL_RECIPIENTS_REPORT = ['goela.shubh@gmail.com', 'goela.engineers@gmail.com']
GOOGLE_SCRIPT_URL = os.getenv('GOOGLE_SCRIPT_URL')

BHAV_COPY_ARCHIVE_PATH = os.getenv('BHAV_COPY_ARCHIVE_PATH', 'bhav_copy_archive')
FO_BHAV_COPY_ARCHIVE_PATH = os.getenv('FO_BHAV_COPY_ARCHIVE_PATH', 'fo_bhav_copy_archive')

INDEX_SYMBOLS = {'NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'SMALLCAPNIFTY', 'NIFTYNXT50'}


def check_for_files( eq_section_id= "cr_equity_daily_Current", der_section_id="cr_deriv_equity_daily_Current"):

    filtered_docs = []
    print(eq_section_id,',',der_section_id)
    # cr_deriv_equity_daily_Previous, cr_deriv_equity_daily_Current
    derivative_file_details = get_display_and_file_names(NSE_base_url, endpoint = '/all-reports-derivatives', section_id = der_section_id)
    if derivative_file_details != []:
        for file in derivative_file_details:
            if file['display_name'] in BHAVCOPY_FILE_NAMES:
                filtered_docs.append(file)
    
    # cr_equity_daily_Previous, cr_equity_daily_Current
    equity_file_details = get_display_and_file_names(NSE_base_url, endpoint = '/all-reports', section_id= eq_section_id)
    if equity_file_details != []:
        for file in equity_file_details:
            if file['display_name'] in BHAVCOPY_FILE_NAMES:
                filtered_docs.append(file)
    
    file_names = [file.get('display_name') for file in filtered_docs]

    return file_names, filtered_docs


def process_filter_docs_for_noti(filtered_docs, template_name = 'bhavcopy_noti', sent_files = []):
    print('in process_filter_docs_for_noti')
    print(sent_files)
    print(filtered_docs)
    if len(filtered_docs) > 0:
        for file in filtered_docs:
            print('generating html...')
            # template = get_mail_template(template_name)
            template = None
            body = generate_html_table(template, [file])

            if file.get('display_name') not in sent_files:
                # file_names.append(file.get('display_name'))
                if template is None:
                    mail_sent = send_email( 
                        recipient_emails=MAIL_RECIPIENTS,
                        # recipient_emails= ['shubh.goela@mnclgroup.com'],
                        subject=file.get('display_name'),
                        body=body,
                        html_body=body)  
                else:  
                    mail_sent = send_email( 
                        recipient_emails=template.get('recipients', MAIL_RECIPIENTS),
                        subject=file.get('display_name'),
                        body=body,
                        html_body=body)
    return True


def extract_date_from_csv(text, year):
    match_1 = re.search(r'([A-Za-z]+)\s(\d{1,2})', text)
    if match_1:
        month = match_1.group(1)
        day = match_1.group(2)
    
    match_2 = re.search(r'\d{4}', year)
    if match_2:
        year = int(match_2.group(0))  # Convert to integer
        
    date_str = f"{day} {month} {year}"
    date_obj = datetime.strptime(date_str, "%d %b %Y").strftime('%d/%m/%y')
    
    return date_obj


def process_open_interest_df(df):
    dt = extract_date_from_csv(df.columns[0], df.columns[1])
    df.dropna(axis=1, how='all', inplace=True)
    df.columns = [v.strip() for v in df.iloc[0]]
    df = df.loc[1:]
    for column in df.columns:
        if df[column].apply(pd.to_numeric, errors='coerce').notna().all():
            df[column] = pd.to_numeric(df[column], errors='coerce')
    
    return df, dt


def create_index_futures_table(cr_df, prev_df, cr_dt, prev_dt):
    merged_df = cr_df.merge(prev_df, on="Client Type", suffixes=(f'_{cr_dt}', f'_{prev_dt}'))
    merged_df["Future Index Long Change"] = merged_df[f"Future Index Long_{cr_dt}"] - merged_df[f"Future Index Long_{prev_dt}"]
    merged_df["Future Index Short Change"] = merged_df[f"Future Index Short_{cr_dt}"] - merged_df[f"Future Index Short_{prev_dt}"]
    merged_df['Signal'] = np.where(
    (merged_df['Future Index Long Change'] - merged_df['Future Index Short Change']) > 0,
        'Bullish',
        'Bearish'
    )
    result_df = merged_df[[
        "Client Type", 
        f"Future Index Long_{cr_dt}", f"Future Index Short_{cr_dt}",  # Current values
        f"Future Index Long_{prev_dt}", f"Future Index Short_{prev_dt}", 
        "Future Index Long Change", "Future Index Short Change", 'Signal' # Daily change
    ]]
    result_df.iloc[-1, -1] = ''
    v1 = result_df.loc[result_df['Client Type'] == 'FII',f'Future Index Long_{cr_dt}'].values[0]
    v2 = result_df.loc[result_df['Client Type'] == 'FII',f'Future Index Short_{cr_dt}'].values[0]
    long_exposure = round(((v1/(v1+v2))*100),2)
    return result_df , long_exposure 


def create_index_call_table(cr_df, prev_df, cr_dt, prev_dt):
    merged_df = cr_df.merge(prev_df, on="Client Type", suffixes=(f'_{cr_dt}', f'_{prev_dt}'))

    merged_df["Option Index Call Long Change"] = merged_df[f"Option Index Call Long_{cr_dt}"] - merged_df[f"Option Index Call Long_{prev_dt}"]
    merged_df["Option Index Call Short Change"] = merged_df[f"Option Index Call Short_{cr_dt}"] - merged_df[f"Option Index Call Short_{prev_dt}"]

    merged_df['Signal'] = np.where(
    (merged_df['Option Index Call Long Change'] - merged_df['Option Index Call Short Change']) > 0,
        'Bullish',
        'Bearish'
    )
    result_df = merged_df[[
        "Client Type", 
        f"Option Index Call Long_{cr_dt}", f"Option Index Call Short_{cr_dt}",  # Current values
        f"Option Index Call Long_{prev_dt}", f"Option Index Call Short_{prev_dt}", 
        "Option Index Call Long Change", "Option Index Call Short Change", 'Signal' # Daily change
    ]]
    result_df.iloc[-1, -1] = ''
    return result_df


def create_index_put_table(cr_df, prev_df, cr_dt, prev_dt):
    merged_df = cr_df.merge(prev_df, on="Client Type", suffixes=(f'_{cr_dt}', f'_{prev_dt}'))

    merged_df["Option Index Put Long Change"] = merged_df[f"Option Index Put Long_{cr_dt}"] - merged_df[f"Option Index Put Long_{prev_dt}"]
    merged_df["Option Index Put Short Change"] = merged_df[f"Option Index Put Short_{cr_dt}"] - merged_df[f"Option Index Put Short_{prev_dt}"]

    merged_df['Signal'] = np.where(
    (merged_df['Option Index Put Long Change'] - merged_df['Option Index Put Short Change']) < 0,
        'Bullish',
        'Bearish'
    )
    result_df = merged_df[[
        "Client Type", 
        f"Option Index Put Long_{cr_dt}", f"Option Index Put Short_{cr_dt}",  # Current values
        f"Option Index Put Long_{prev_dt}", f"Option Index Put Short_{prev_dt}", 
        "Option Index Put Long Change", "Option Index Put Short Change", 'Signal' # Daily change
    ]]
    result_df.iloc[-1, -1] = ''
    return result_df


def create_stock_futures_table(cr_df, prev_df, cr_dt, prev_dt):
    merged_df = cr_df.merge(prev_df, on="Client Type", suffixes=(f'_{cr_dt}', f'_{prev_dt}'))
    merged_df["Future Stock Long Change"] = merged_df[f"Future Stock Long_{cr_dt}"] - merged_df[f"Future Stock Long_{prev_dt}"]
    merged_df["Future Stock Short Change"] = merged_df[f"Future Stock Short_{cr_dt}"] - merged_df[f"Future Stock Short_{prev_dt}"]
    merged_df['Signal'] = np.where(
    (merged_df['Future Stock Long Change'] - merged_df['Future Stock Short Change']) > 0,
        'Bullish',
        'Bearish'
    )
    result_df = merged_df[[
        "Client Type", 
        f"Future Stock Long_{cr_dt}", f"Future Stock Short_{cr_dt}",  # Current values
        f"Future Stock Long_{prev_dt}", f"Future Stock Short_{prev_dt}", 
        "Future Stock Long Change", "Future Stock Short Change", 'Signal' # Daily change
    ]]
    result_df.iloc[-1, -1] = ''
    v1 = result_df.loc[result_df['Client Type'] == 'FII',f'Future Stock Long_{cr_dt}'].values[0]
    v2 = result_df.loc[result_df['Client Type'] == 'FII',f'Future Stock Short_{cr_dt}'].values[0]
    long_exposure = round(((v1/(v1+v2))*100),2)
    return result_df , long_exposure 


def create_stock_call_table(cr_df, prev_df, cr_dt, prev_dt):
    merged_df = cr_df.merge(prev_df, on="Client Type", suffixes=(f'_{cr_dt}', f'_{prev_dt}'))

    merged_df["Option Stock Call Long Change"] = merged_df[f"Option Stock Call Long_{cr_dt}"] - merged_df[f"Option Stock Call Long_{prev_dt}"]
    merged_df["Option Stock Call Short Change"] = merged_df[f"Option Stock Call Short_{cr_dt}"] - merged_df[f"Option Stock Call Short_{prev_dt}"]

    merged_df['Signal'] = np.where(
    (merged_df['Option Stock Call Long Change'] - merged_df['Option Stock Call Short Change']) > 0,
        'Bullish',
        'Bearish'
    )
    result_df = merged_df[[
        "Client Type", 
        f"Option Stock Call Long_{cr_dt}", f"Option Stock Call Short_{cr_dt}",  # Current values
        f"Option Stock Call Long_{prev_dt}", f"Option Stock Call Short_{prev_dt}", 
        "Option Stock Call Long Change", "Option Stock Call Short Change", 'Signal' # Daily change
    ]]
    result_df.iloc[-1, -1] = ''
    return result_df


def create_stock_put_table(cr_df, prev_df, cr_dt, prev_dt):
    merged_df = cr_df.merge(prev_df, on="Client Type", suffixes=(f'_{cr_dt}', f'_{prev_dt}'))

    merged_df["Option Stock Put Long Change"] = merged_df[f"Option Stock Put Long_{cr_dt}"] - merged_df[f"Option Stock Put Long_{prev_dt}"]
    merged_df["Option Stock Put Short Change"] = merged_df[f"Option Stock Put Short_{cr_dt}"] - merged_df[f"Option Stock Put Short_{prev_dt}"]

    merged_df['Signal'] = np.where(
    (merged_df['Option Stock Put Long Change'] - merged_df['Option Stock Put Short Change']) < 0,
        'Bullish',
        'Bearish'
    )
    result_df = merged_df[[
        "Client Type", 
        f"Option Stock Put Long_{cr_dt}", f"Option Stock Put Short_{cr_dt}",  # Current values
        f"Option Stock Put Long_{prev_dt}", f"Option Stock Put Short_{prev_dt}", 
        "Option Stock Put Long Change", "Option Stock Put Short Change", 'Signal' # Daily change
    ]]
    result_df.iloc[-1, -1] = ''
    return result_df


def create_OI_table(fo_bhav_copy, XpryDts, index):
    df = pd.DataFrame()
    df[index] = ['Max Call OI','Max Put OI', 'Change in Call OI max', 'Change in Put OI max']
    df['Option Type'] = ['CE','PE','CE','PE']

    for date in XpryDts:
        
        nifty_call = fo_bhav_copy[(fo_bhav_copy['XpryDt'] == date.strftime('%Y-%m-%d')) & (fo_bhav_copy['TckrSymb'] == index)  & (fo_bhav_copy['OptnTp'] == 'CE')]
        nifty_call = nifty_call.loc[nifty_call['OpnIntrst'].idxmax()]['StrkPric']


        nifty_put = fo_bhav_copy[(fo_bhav_copy['XpryDt'] == date.strftime('%Y-%m-%d')) & (fo_bhav_copy['TckrSymb'] == index)  & (fo_bhav_copy['OptnTp'] == 'PE')]
        nifty_put = nifty_put.loc[nifty_put['OpnIntrst'].idxmax()]['StrkPric']

        nifty_call_max_oi_chng = fo_bhav_copy[(fo_bhav_copy['XpryDt'] == date.strftime('%Y-%m-%d')) & (fo_bhav_copy['TckrSymb'] == index)  & (fo_bhav_copy['OptnTp'] == 'CE')]
        nifty_call_max_oi_chng = nifty_call_max_oi_chng.loc[nifty_call_max_oi_chng['ChngInOpnIntrst'].idxmax()]['StrkPric']

        nifty_put_max_oi_chng = fo_bhav_copy[(fo_bhav_copy['XpryDt'] == date.strftime('%Y-%m-%d')) & (fo_bhav_copy['TckrSymb'] == index)  & (fo_bhav_copy['OptnTp'] == 'PE')]
        nifty_put_max_oi_chng = nifty_put_max_oi_chng.loc[nifty_put_max_oi_chng['ChngInOpnIntrst'].idxmax()]['StrkPric']

        df[date.strftime('%d-%b')] = [nifty_call, nifty_put, nifty_call_max_oi_chng, nifty_put_max_oi_chng]

    return df


def get_pcr(fo_bhav_copy, XpryDt, index, ref_date = None):
    if ref_date is None:
        ref_date = datetime.today()
    elif isinstance(ref_date, str):
        ref_date = datetime.strptime(ref_date, "%Y-%m-%d")

    if ref_date.date() == XpryDt:
        fo_data_filterd = fo_bhav_copy[(fo_bhav_copy['XpryDt'] != ref_date.strftime('%Y-%m-%d')) & (fo_bhav_copy['TckrSymb'] == index)]
    else:
        fo_data_filterd = fo_bhav_copy[(fo_bhav_copy['TckrSymb'] == index)]

    fo_data_filterd_ce = fo_data_filterd[(fo_data_filterd['OptnTp'] == 'CE')]
    fo_data_filterd_pe = fo_data_filterd[(fo_data_filterd['OptnTp'] == 'PE')]

    # return fo_data_filterd, fo_data_filterd_ce, fo_data_filterd_pe
    return fo_data_filterd_pe['OpnIntrst'].sum()/ fo_data_filterd_ce['OpnIntrst'].sum()


def create_stock_analysis_report(curr_bhav, fo_bhav_copy, month_expry, symbols):
    """
    Create a stock analysis report using the FinancialDataAnalyzer class.
    
    Args:
        fo_bhav_copy: DataFrame with futures and options data
        symbol: Trading symbol (e.g., 'NIFTY')
        month_expry: Current month expiry date
        curr_close: Current closing price
        prev_close: Previous closing price
    
    Returns:
        HTML report string
    """


    html = ''' '''
    for symbol in symbols:
        prev_close = curr_bhav[curr_bhav['SYMBOL'] == symbol][' PREV_CLOSE'].values[0]
        curr_close = curr_bhav[curr_bhav['SYMBOL'] == symbol][' CLOSE_PRICE'].values[0]
        analyzer = FinancialDataAnalyzer(fo_bhav_copy, symbol, month_expry, curr_close, prev_close)
        html += analyzer.get_html_report() + '<br><hr>'
    
    html += '<p style="margin-top:20px; font-size:12px; color:#6b7280;">Generated automatically</p>'
    return html


def create_complete_OI_data_from_bhavcopy(curr_bhav, curr_fo_bhav, month_expry):

    fo_bhav_symbols = set(curr_fo_bhav['TckrSymb'].unique())
    symbols = fo_bhav_symbols - INDEX_SYMBOLS

    summary_rows = []

    for symbol in symbols:
        print(f"Generating summary for {symbol}...")
        prev_close = curr_bhav[curr_bhav['SYMBOL'] == symbol][' PREV_CLOSE'].values[0]
        curr_close = curr_bhav[curr_bhav['SYMBOL'] == symbol][' CLOSE_PRICE'].values[0]
        analyzer = FinancialDataAnalyzer(curr_fo_bhav, symbol, month_expry, curr_close, prev_close)


        # ATM strike = nearest strike to current price
        atm_strikes_data = analyzer.analyze_options_activity()["strikes_data"]
        strikes = atm_strikes_data.keys()
        atm_strike = min(strikes, key=lambda x: abs(x - curr_close))

        # Max OI and additions/unwinding
        oi_positions = analyzer.analyze_max_oi_positions()
        
        # PCR + totals
        put_oi, call_oi, pcr, total_oi, chng_in_oi, pct_chng_in_oi = analyzer.get_pcr_and_total_oi()

        summary = {

            # Equity Details
            "Symbol": symbol,
            "Prev Close": prev_close,
            "Curr Close": curr_close,
            "Price Change": curr_close - prev_close,
            "Price Change %": round(((curr_close - prev_close) / prev_close) * 100, 2) if prev_close != 0 else 0,
            "PCR": pcr,

            # ATM Strike details
            "ATM Strike": atm_strike,
            "ATM Call Price": atm_strikes_data[atm_strike].get("CE", {}).get("settlement_price", ''),
            "ATM Call Price %Chg": atm_strikes_data[atm_strike].get("CE", {}).get("price_change_pct", ''),
            "ATM Call OI": atm_strikes_data[atm_strike].get("CE", {}).get("oi", ''),
            "ATM Call OI %Chg": atm_strikes_data[atm_strike].get("CE", {}).get("oi_change_pct", ''),
            
            "ATM Put Price": atm_strikes_data[atm_strike].get("PE", {}).get("settlement_price", ''),
            "ATM Put Price %Chg": atm_strikes_data[atm_strike].get("PE", {}).get("price_change_pct", ''),
            "ATM Put OI": atm_strikes_data[atm_strike].get("PE", {}).get("oi", ''),
            "ATM Put OI %Chg": atm_strikes_data[atm_strike].get("PE", {}).get("oi_change_pct", ''),

            "ATM Total OI": atm_strikes_data[atm_strike].get("CE", {}).get("oi", 0) + atm_strikes_data[atm_strike].get("PE", {}).get("oi", 0),
            "ATM Total OI %Chg": round(((atm_strikes_data[atm_strike].get("CE", {}).get("oi_change", 0) + 
                                atm_strikes_data[atm_strike].get("PE", {}).get("oi_change", 0)) /

                                (atm_strikes_data[atm_strike].get("CE", {}).get("oi", 0) + 
                                atm_strikes_data[atm_strike].get("PE", {}).get("oi", 0) - 
                                (atm_strikes_data[atm_strike].get("CE", {}).get("oi_change", 0) + 
                                atm_strikes_data[atm_strike].get("PE", {}).get("oi_change", 0)
                                )
                                )
                                ) * 100, 2)
                                if (atm_strikes_data[atm_strike].get("CE", {}).get("oi", 0) + 
                                    atm_strikes_data[atm_strike].get("PE", {}).get("oi", 0) - 
                                    (atm_strikes_data[atm_strike].get("CE", {}).get("oi_change", 0) + 
                                        atm_strikes_data[atm_strike].get("PE", {}).get("oi_change", 0))) != 0 else 0, 


            # OI Totals
            "Total Put OI": put_oi,
            "Total Call OI": call_oi,
            "Total OI": total_oi,
            "Change in Total OI": chng_in_oi,
            "Change in Total OI %": pct_chng_in_oi,

            # CE OI Positions
            "Max Call OI Strike": oi_positions["Max Call OI"]["strike"],
            "Max Call OI": oi_positions["Max Call OI"]["oi"],
            "Max Call OI %Chg": oi_positions["Max Call OI"]["oi_change_pct"],
            "Max Call OI Add Strike": oi_positions["Max Call OI addition"]["strike"],
            "Max Call OI Add": oi_positions["Max Call OI addition"]["oi_change"],
            "Max Call OI Add %Chg": oi_positions["Max Call OI addition"]["oi_change_pct"],
            "Max Call OI Unwind Strike": oi_positions["Max Call OI unwinding"]["strike"],
            "Max Call OI Unwind": oi_positions["Max Call OI unwinding"]["oi_change"],
            "Max Call OI Unwind %Chg": oi_positions["Max Call OI unwinding"]["oi_change_pct"],

            # PE OI Positions
            "Max Put OI Strike": oi_positions["Max Put OI"]["strike"],
            "Max Put OI": oi_positions["Max Put OI"]["oi"],
            "Max Put OI %Chg": oi_positions["Max Put OI"]["oi_change_pct"],
            "Max Put OI Add Strike": oi_positions["Max Put OI addition"]["strike"],
            "Max Put OI Add": oi_positions["Max Put OI addition"]["oi_change"],
            "Max Put OI Add %Chg": oi_positions["Max Put OI addition"]["oi_change_pct"],
            "Max Put OI Unwind Strike": oi_positions["Max Put OI unwinding"]["strike"],
            "Max Put OI Unwind": oi_positions["Max Put OI unwinding"]["oi_change"],
            "Max Put OI Unwind %Chg": oi_positions["Max Put OI unwinding"]["oi_change_pct"]
        }

        summary_rows.append(summary)

    return summary_rows


def create_and_archive_OI_excel_report(summary_rows):

    df = pd.DataFrame(summary_rows)
    df.sort_values(by=["Change in Total OI %","ATM Total OI %Chg"], ascending=False, inplace=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Options Summary"

    # Thin border style for section outlines
    thin = Side(border_style="thin", color="000000")

    # Define top-level sections (start_col, end_col)
    sections = [
        ("Equity Price", 1, 6),
        ("ATM", 7, 17),
        ("OI Totals", 18, 22),
        ("Call Options", 23, 31),
        ("Put Options", 32, 40)
    ]

    # Top header row (merged)
    for title, start, end in sections:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply border only at section edges
        for col in range(start, end + 1):
            cell = ws.cell(row=1, column=col)
            if col == start:
                cell.border = Border(left=thin, top=thin, bottom=thin)
            elif col == end:
                cell.border = Border(right=thin, top=thin, bottom=thin)
            else:
                cell.border = Border(top=thin, bottom=thin)
            cell.fill = PatternFill("solid", fgColor="BDD7EE")

    # Second header row (column names)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="FFD966")

    # Write data rows with alternate coloring (no inner borders)
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        fill_color = "FFF2CC" if r_idx % 2 == 1 else "FFFFFF"
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=fill_color)

    # Apply outer borders for each section on data rows
    max_row = ws.max_row
    for _, start, end in sections:
        # top border
        for col in range(start, end + 1):
            ws.cell(row=2, column=col).border = Border(top=thin)
        # bottom border
        for col in range(start, end + 1):
            ws.cell(row=max_row, column=col).border = Border(bottom=thin)
        # left & right borders
        for row in range(2, max_row + 1):
            ws.cell(row=row, column=start).border = Border(left=thin)
            ws.cell(row=row, column=end).border = Border(right=thin)

    # Adjust column widths
    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = max(len(col)+2, 12)


    try:
        file_path = f"Data/OI_data/OI_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(file_path)
    except Exception as e:
        print(f"Error saving Excel file: {e}")


    return file_path


def loop_question_between_times(start_time="00:00", end_time="23:00", interval=60):
    """
    Continuously prompts a question between specified times.

    Parameters:
        question (str): The question to be asked in the loop.
        start_time (str): The start time in HH:MM format (24-hour format).
        end_time (str): The end time in HH:MM format (24-hour format).
        interval (int): The number of seconds to wait between each prompt.

    """
    today = datetime.now().strftime('%Y-%m-%d')
    start = datetime.strptime(start_time, "%H:%M").time()
    end = datetime.strptime(end_time, "%H:%M").time()

    session = login(NSE_base_url)
    MAX_RETRIES = 5 
    retry_count = 0
    current_df, prev_df, curr_fo_bhav, curr_bhav = None, None, None, None

    sent_files = []
    filtered_docs = []
    while True:
        now = datetime.now().time()
        # print('now: ', now)
        # print(files)
        if start <= now <= end and set(sent_files) != set(BHAVCOPY_FILE_NAMES):
            f, fd = check_for_files( eq_section_id="cr_equity_daily_Current" , der_section_id="cr_deriv_equity_daily_Current")
            process_filter_docs_for_noti(filtered_docs = fd, sent_files=sent_files)
            if f != []:
                sent_files.extend(f)
                sent_files = list(set(sent_files))
                filtered_docs.extend(fd)
                filtered_docs = list({tuple(sorted(d.items())) for d in filtered_docs})
                filtered_docs = [dict(t) for t in filtered_docs]

            print('sleeping...')
            time.sleep(interval)
        else:
            if now > end:
                print("The time window has closed. Exiting loop.")
            break
    


    while retry_count <= MAX_RETRIES:
        retry_count += 1
        for file in filtered_docs:
            if file['display_name'] == 'F&O-Participant wise Open Interest (csv)':
                response = session.get(file['file_link'])
                try:
                    response.raise_for_status()
                    if response.status_code == 200:
                        current_df = handle_file_response(response, today)
                        current_df, current_date = process_open_interest_df(current_df)
                except Exception as e:
                    current_df = None
                    print(f"Error processing current OI file: {e}")
                    continue
            
            if file['display_name'] == 'F&O-UDiFF Common Bhavcopy Final (zip)':
                response = session.get(file['file_link'])
                try:
                    response.raise_for_status()
                    if response.status_code == 200:
                        curr_fo_bhav = handle_file_response(response, today)
                except Exception as e:
                    curr_fo_bhav = None
                    print(f"Error processing current F&O Bhavcopy file: {e}")
                    continue
            
            if file['display_name'] == 'Full Bhavcopy and Security Deliverable data':
                response = session.get(file['file_link'])
                try:
                    response.raise_for_status()
                    if response.status_code == 200:
                        curr_bhav = handle_file_response(response, today)
                except Exception as e:
                    curr_bhav = None
                    print(f"Error processing current F&O Bhavcopy file: {e}")
                    continue

            if current_df is not None and curr_fo_bhav is not None and curr_bhav is not None:
                break
            

        f, fd = check_for_files( eq_section_id="cr_equity_daily_Previous" , der_section_id="cr_deriv_equity_daily_Previous")
        for file in fd:
            if file['display_name'] == 'F&O-Participant wise Open Interest (csv)':
                response = session.get(file['file_link'])
                try:
                    response.raise_for_status()
                    if response.status_code == 200:
                        prev_df = handle_file_response(response, today)
                        prev_df, prev_date = process_open_interest_df(prev_df)
                        break
                except Exception as e:
                    prev_df = None
                    print(f"Error processing previous OI file: {e}")
                    continue
        
        if current_df is not None and prev_df is not None:
            break


    print('########curren_df')
    print(current_df)
    print('########previous_df')
    print(prev_df)


    r1, index_long_exposure = create_index_futures_table(current_df, prev_df, current_date, prev_date)
    r2 = create_index_call_table(current_df, prev_df, current_date, prev_date)
    r3 = create_index_put_table(current_df, prev_df, current_date, prev_date)
    r4, stock_long_exposure = create_stock_futures_table(current_df, prev_df, current_date, prev_date)
    r5 = create_stock_call_table(current_df, prev_df, current_date, prev_date)
    r6 = create_stock_put_table(current_df, prev_df, current_date, prev_date)

    # month_expry_day = get_expry_days(exchange='NSE', index='NIFTY', expiry_type='month')
    # week_expry_day = get_expry_days(exchange='NSE', index='NIFTY', expiry_type='week')

    month_expry_day = week_expry_day = 'Tuesday'
    week_expry, month_expry = get_expiry_dates(weekly_day= week_expry_day, monthly_day=month_expry_day)

    OI_table = create_OI_table(curr_fo_bhav, [week_expry, month_expry], 'NIFTY')
    nifty_pcr = get_pcr(curr_fo_bhav, XpryDt = week_expry, index = 'NIFTY' )


    extra_table_data = create_html_for_exposure(index_long_exposure, stock_long_exposure)
    commentry = get_commentry(r1, r2, r3, index_long_exposure, nifty_pcr, OI_table, week_expry, month_expry)
    html = create_html_table_with_predefined_html([{'heading': 'Index Futures', 'df': r1},
                                                {'heading': 'Stock Futures', 'df': r4},
                                                {'heading': 'Index Call Options', 'df': r2},
                                                {'heading': 'Stock Call Options', 'df': r5},
                                                {'heading': 'Index Put Options', 'df': r3},
                                                {'heading': 'Stock Put Options', 'df': r6}], extra_table_data, commentry)
    

    try:
        save_html_as_png(html_string=html)
    except Exception as e:
        print('Failed to save html as image')
        pass

    send_email( 
                recipient_emails=MAIL_RECIPIENTS_REPORT,
                subject='Participant Wise Derivatives FII-DII Data',
                body=html,
                html_body=html,
                attachments=['output.png'])

    delete_file('output.png')

    curr_bhav_filepath = save_bhav_copy_data(bhav_copy_archive_path=BHAV_COPY_ARCHIVE_PATH, date=datetime.now(), data=curr_bhav)
    curr_bhav_fo_filepath = save_bhav_copy_data(bhav_copy_archive_path=FO_BHAV_COPY_ARCHIVE_PATH, date=datetime.now(), data=curr_fo_bhav)

    create_or_add_master_data(curr_bhav_filepath, today, "open")
    create_or_add_master_data(curr_bhav_filepath, today, "high")
    create_or_add_master_data(curr_bhav_filepath, today, "low")
    create_or_add_master_data(curr_bhav_filepath, today, "close")
    create_or_add_master_data(curr_bhav_filepath, today, "volume")
    create_or_add_master_data(curr_bhav_filepath, today, "delivery_qty")
    create_or_add_master_data(curr_bhav_filepath, today, "delivery_per")

    adjust_corp_actions(date=datetime.now())

    try:
        df = get_google_script_data(GOOGLE_SCRIPT_URL)
        symbols = list(df['STOCK_NAME'])
        html = create_stock_analysis_report(curr_bhav, curr_fo_bhav, month_expry, symbols)
    except Exception as e:
        print(f"Failed to generate stock analysis report: {e}")
        html = '''<p style="color:red;">Failed to generate stock analysis report</p>'''
        pass
    
    try:
        summary_rows = create_complete_OI_data_from_bhavcopy(curr_bhav, curr_fo_bhav, month_expry)
    except Exception as e:
        print(e)
        summary_rows = []
        pass

    file_path = create_and_archive_OI_excel_report(summary_rows)

    send_email( 
        recipient_emails=MAIL_RECIPIENTS_REPORT,
        subject='Stock Analysis Report',
        body=html,
        html_body=html,
        attachments=[file_path])

    return True


def daily_runner():
    last_processed_date = None

    while True:
        now = datetime.now()
        today_str = now.strftime('%Y-%m-%d')
        start_time = now.replace(hour=18, minute=0, second=0, microsecond=0)
        end_time = now.replace(hour=23, minute=59, second=59, microsecond=0)

        if start_time <= now <= end_time and today_str != last_processed_date and is_valid_date(now):
            print(f"Running process for {today_str}")
            try:
                result = loop_question_between_times()
                if result:
                    last_processed_date = today_str
            except Exception as e:
                print(f"Error processing data for {today_str}: {e}")
        else:
            print(f"Waiting... ({now.strftime('%H:%M:%S')})")
        
        next_run_time = get_next_valid_trading_datetime(now)
        sleep_duration = (next_run_time - datetime.now()).total_seconds()
        print(f"Sleeping for {int(sleep_duration)} seconds until {next_run_time.strftime('%Y-%m-%d %H:%M:%S')}")
        time.sleep(sleep_duration)


if __name__ == "__main__":
    daily_runner()
